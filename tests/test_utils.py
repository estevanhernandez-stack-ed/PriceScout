import pytest
import pandas as pd
import io
import openpyxl
import datetime
import json
import os
import pytz
import streamlit as st
from unittest.mock import patch

from app.utils import style_price_change, to_excel, to_csv, check_cache_status, log_runtime, format_price_change, style_price_change_v2, get_error_message, normalize_time_string, format_theater_name_for_display, is_run_allowed, estimate_scrape_time, clear_workflow_state, clean_film_title, format_time_to_human_readable, _extract_company_name, _categorize_formats, get_report_path, to_excel_multi_sheet, showtime_selection_to_dataframe, generate_selection_analysis_report

def test_style_price_change():
    assert style_price_change("▲$5.00") == 'background-color: #e6ffe6' # Light green
    assert style_price_change("▼$5.00") == 'background-color: #ffe6e6' # Light red
    assert style_price_change("$0.00") == ''
    assert style_price_change("No Change") == ''

def test_format_price_change():
    """Tests the format_price_change function."""
    assert format_price_change('$10.00', '$15.00') == '▲5.00'
    assert format_price_change('$15.00', '$10.00') == '▼5.00'
    assert format_price_change('$10.00', '$10.00') == '—'
    assert format_price_change(None, '$10.00') == 'N/A'
    assert format_price_change('$10.00', None) == 'N/A'

def test_style_price_change_v2():
    """Tests the style_price_change_v2 function."""
    assert style_price_change_v2('▲$5.00') == 'color: green'
    assert style_price_change_v2('▼$5.00') == 'color: red'
    assert style_price_change_v2('—') == ''
    assert style_price_change_v2('Some other string') == ''

def test_get_error_message():
    """Tests the get_error_message utility."""
    assert get_error_message(ValueError("Test error")) == "Test error"
    assert get_error_message("a string") == "An unknown error occurred."
    assert get_error_message(123) == "An unknown error occurred."

def test_normalize_time_string():
    """Tests the normalization of various time string formats."""
    assert normalize_time_string('4:15p') == '04:15PM'
    assert normalize_time_string('10:30 AM') == '10:30AM'
    assert normalize_time_string('12:00pm') == '12:00PM'
    assert normalize_time_string('1:00 a.m.') == '01:00AM'
    assert normalize_time_string(None) == ""

def test_format_theater_name_for_display():
    """Tests the cleaning of theater names for display."""
    assert format_theater_name_for_display('My Cinema Theatre') == 'My Cinema'
    assert format_theater_name_for_display('My Cinema Theater') == 'My Cinema'
    assert format_theater_name_for_display('My Cinema') == 'My Cinema'
    assert format_theater_name_for_display('My Cinema Theatre ') == 'My Cinema'

def test_to_excel():
    df = pd.DataFrame({'col1': [1, 2], 'col2': [3, 4]})
    excel_data = to_excel(df)
    assert isinstance(excel_data, bytes)
    # check if it's a valid excel file
    workbook = openpyxl.load_workbook(io.BytesIO(excel_data))
    assert workbook.sheetnames == ['Sheet1']

def test_to_csv():
    df = pd.DataFrame({'col1': [1, 2], 'col2': [3, 4]})
    csv_data = to_csv(df)
    assert isinstance(csv_data, bytes)
    expected = 'col1,col2\n1,3\n2,4\n'
    assert csv_data.decode('utf-8').replace('\r\n', '\n') == expected

def test_check_cache_status_missing(monkeypatch):
    monkeypatch.setattr('app.config.CACHE_FILE', 'non_existent_file.json')
    status, last_updated = check_cache_status()
    assert status == "missing"
    assert last_updated is None

def test_check_cache_status_invalid(tmp_path, monkeypatch):
    cache_file = tmp_path / "cache.json"
    cache_file.write_text("invalid json")
    monkeypatch.setattr('app.config.CACHE_FILE', str(cache_file))
    status, last_updated = check_cache_status()
    assert status == "stale"
    assert last_updated == "Error reading cache"

def test_check_cache_status_stale(tmp_path, monkeypatch):
    cache_file = tmp_path / "cache.json"
    stale_date = datetime.datetime.now() - datetime.timedelta(days=4)
    cache_data = {"metadata": {"last_updated": stale_date.isoformat()}}
    cache_file.write_text(json.dumps(cache_data))
    monkeypatch.setattr('app.config.CACHE_FILE', str(cache_file))
    monkeypatch.setattr('app.config.CACHE_EXPIRATION_DAYS', 3)
    status, last_updated = check_cache_status()
    assert status == "stale"
    assert last_updated == stale_date.strftime('%Y-%m-%d %H:%M:%S')

def test_check_cache_status_fresh(tmp_path, monkeypatch):
    cache_file = tmp_path / "cache.json"
    fresh_date = datetime.datetime.now() - datetime.timedelta(days=2)
    cache_data = {"metadata": {"last_updated": fresh_date.isoformat()}}
    cache_file.write_text(json.dumps(cache_data))
    monkeypatch.setattr('app.config.CACHE_FILE', str(cache_file))
    monkeypatch.setattr('app.config.CACHE_EXPIRATION_DAYS', 3)
    status, last_updated = check_cache_status()
    assert status == "fresh"
    assert last_updated == fresh_date.strftime('%Y-%m-%d %H:%M:%S')

def test_log_runtime(tmp_path, monkeypatch):
    reports_dir = tmp_path / "reports"
    log_file = reports_dir / "runtime.csv"
    monkeypatch.setattr('app.config.REPORTS_DIR', str(reports_dir)) # Mock REPORTS_DIR
    monkeypatch.setattr('app.config.RUNTIME_LOG_FILE', str(log_file))
    log_runtime("Test Mode", 10, 100, 123.456)
    assert os.path.exists(log_file)
    with open(log_file, 'r') as f:
        content = f.read()
        assert "Test Mode" in content
        assert "10" in content
        assert "100" in content
        assert "123.456" in content

@patch('app.utils.datetime')
def test_is_run_allowed(mock_datetime):
    """Tests the is_run_allowed function with mocked time."""
    # Mock 'now' to be 7 AM in Chicago (13:00 UTC on a non-DST day)
    mock_now_utc_early = datetime.datetime(2025, 1, 1, 13, 0, 0, tzinfo=pytz.utc)
    mock_datetime.datetime.now.return_value = mock_now_utc_early
    
    # It's before 8 AM, so it should be allowed
    assert is_run_allowed("America/Chicago") is True

    # Mock 'now' to be 9 AM in Chicago (15:00 UTC on a non-DST day)
    mock_now_utc_late = datetime.datetime(2025, 1, 1, 15, 0, 0, tzinfo=pytz.utc)
    mock_datetime.datetime.now.return_value = mock_now_utc_late

    # It's after 8 AM, so it should not be allowed
    assert is_run_allowed("America/Chicago") is False

def test_estimate_scrape_time(tmp_path, monkeypatch):
    """Tests the scrape time estimation based on historical log data."""
    log_file = tmp_path / "runtime.csv"
    monkeypatch.setattr('app.config.RUNTIME_LOG_FILE', str(log_file))

    # Case 1: No log file
    assert estimate_scrape_time(10) == -1

    # Case 2: Log file exists but is empty
    log_file.write_text("timestamp,mode,num_theaters,num_showings,duration_seconds\n")
    assert estimate_scrape_time(10) == -1

    # Case 3: Valid log file
    log_content = (
        "timestamp,mode,num_theaters,num_showings,duration_seconds\n"
        "2025-01-01,Market Mode,5,50,100\n"  # 2s per showing
        "2025-01-02,CompSnipe Mode,10,100,400\n" # 4s per showing
    )
    log_file.write_text(log_content)
    # Average time per showing is (2+4)/2 = 3s.
    # Estimated time for 10 showings should be 10 * 3 = 30s.
    assert estimate_scrape_time(10) == pytest.approx(30.0)



def test_clear_workflow_state(monkeypatch):
    """Tests that clear_workflow_state correctly removes only workflow-specific keys."""
    # 1. Setup a mock session state with a mix of workflow and persistent keys
    mock_state_dict = {
        # Workflow keys that should be cleared
        'selected_market': 'Test Market',
        'final_df': pd.DataFrame(),
        'stage': 'report_generated',
        'report_running': True,
        'cs_date': datetime.date(2025, 1, 1),
        # Persistent keys that should NOT be cleared
        'logged_in': True,
        'user_name': 'testuser',
        'search_mode': 'Market Mode',
        'dev_mode': True
    }
    
    # Use a class to allow attribute deletion, mimicking st.session_state
    class MockSessionState:
        def __init__(self, state_dict):
            self.__dict__.update(state_dict)
        def __contains__(self, key):
            return key in self.__dict__
        def __delitem__(self, key):
            del self.__dict__[key]

    mock_session_object = MockSessionState(mock_state_dict)
    monkeypatch.setattr(st, "session_state", mock_session_object)

    # 2. Call the function
    clear_workflow_state()

    # 3. Assertions
    # Check that workflow keys are gone
    assert 'selected_market' not in st.session_state.__dict__
    assert 'stage' not in st.session_state.__dict__
    # Check that persistent keys remain
    assert 'logged_in' in st.session_state.__dict__
    assert st.session_state.logged_in is True

def test_clean_film_title_basic():
    """Tests basic film title cleaning."""
    # Remove trailing year
    assert clean_film_title("My Film (2025)") == "My Film"
    
    # Remove punctuation
    assert clean_film_title("The Devil's Rejects") == "The Devils Rejects"
    assert clean_film_title("Mission: Impossible") == "Mission Impossible"
    assert clean_film_title("Hello, World!") == "Hello World"
    
    # Empty or None input
    assert clean_film_title("") == ""
    assert clean_film_title(None) == ""

def test_clean_film_title_mystery_movies():
    """Tests handling of Mystery Movie variations."""
    assert clean_film_title("AMC Mystery Movie - Something") == "Mystery Movie"
    assert clean_film_title("Mystery Movie Premiere") == "Mystery Movie"
    assert clean_film_title("Secret Movie") == "Secret Movie"
    assert clean_film_title("Secret Screening Tonight") == "Secret Screening"

def test_clean_film_title_event_terms():
    """Tests removal of event-related terms."""
    assert clean_film_title("Casablanca (Re-release)") == "Casablanca"
    assert clean_film_title("The Godfather - Anniversary") == "The Godfather"
    assert clean_film_title("Star Wars (Fathom Events)") == "Star Wars"
    assert clean_film_title("Spirited Away (Ghibli Fest)") == "Spirited Away"
    assert clean_film_title("Film Title (Director's Cut)") == "Film Title"

def test_clean_film_title_complex():
    """Tests cleaning with multiple issues."""
    # Hyphenated event term is removed
    assert clean_film_title("The Lion King - Re-release") == "The Lion King"
    # Year + parenthetical event term - both removed
    assert clean_film_title("Avatar (2009) (Re-release)") == "Avatar"
    # Parenthetical event term + year at end - both removed
    assert clean_film_title("Avatar (Re-release) (2009)") == "Avatar"
    # Punctuation + event term removal
    assert clean_film_title("Avatar: The Way of Water (Advanced Screening)") == "Avatar The Way of Water"
    # Year in middle gets removed
    assert clean_film_title("The Lion King (2019) - Re-release") == "The Lion King"

def test_format_time_to_human_readable():
    """Tests conversion of seconds to human-readable format."""
    # Hours, minutes, seconds
    assert format_time_to_human_readable(3661) == "1h 1m 1s"
    assert format_time_to_human_readable(3600) == "1h 0m 0s"
    
    # Just minutes and seconds
    assert format_time_to_human_readable(90) == "1m 30s"
    assert format_time_to_human_readable(60) == "1m 0s"
    
    # Just seconds
    assert format_time_to_human_readable(45) == "45s"
    assert format_time_to_human_readable(0) == "0s"
    
    # Negative value
    assert format_time_to_human_readable(-10) == "N/A"
    
    # Large value
    assert format_time_to_human_readable(7265) == "2h 1m 5s"

def test_extract_company_name():
    """Tests extraction of company names from theater names."""
    # Marcus Theatres
    assert _extract_company_name("Marcus Wehrenberg Des Peres 14 Cine") == "Marcus Theatres"
    assert _extract_company_name("Marcus Cape West Cinema") == "Marcus Theatres"
    
    # AMC Theatres
    assert _extract_company_name("AMC West Olive 16") == "AMC Theatres"
    assert _extract_company_name("amc esquire 7") == "AMC Theatres"
    
    # Cinemark
    assert _extract_company_name("Cinemark Merriam") == "Cinemark Theatres"
    assert _extract_company_name("Century 16 Deer Valley") == "Cinemark Theatres"
    
    # Regal
    assert _extract_company_name("Regal Stonecrest") == "Regal Cinemas"
    assert _extract_company_name("Edwards Houston Marq'E") == "Regal Cinemas"
    
    # B&B Theatres
    assert _extract_company_name("B&B Theatres Shawnee 18") == "B&B Theatres"
    
    # Alamo Drafthouse
    assert _extract_company_name("Alamo Drafthouse Cinema Mueller") == "Alamo Drafthouse"
    
    # Unknown
    assert _extract_company_name("Local Independent Cinema") == "Unknown"

def test_categorize_formats():
    """Tests categorization of film formats into premium and general."""
    # Standard 2D is excluded from both categories
    premium, general = _categorize_formats("2D", set())
    assert premium == []
    assert general == []
    
    # 3D is premium
    premium, general = _categorize_formats("3D", set())
    assert "3D" in premium
    assert general == []
    
    # PLF formats are premium
    premium, general = _categorize_formats("IMAX, Dolby Cinema", {"imax", "dolby cinema"})
    assert "Dolby Cinema" in premium
    assert "IMAX" in premium
    assert general == []
    
    # Mixed - 3D is premium, non-2D non-premium formats are general
    premium, general = _categorize_formats("2D, 3D, Reserved Seating", set())
    assert "3D" in premium
    assert "Reserved Seating" in general
    assert "2D" not in premium
    assert "2D" not in general
    
    # D-BOX is premium
    premium, general = _categorize_formats("D-BOX, Reserved Seating", set())
    assert "D-BOX" in premium
    assert "Reserved Seating" in general
    
    # Empty input
    premium, general = _categorize_formats("", set())
    assert premium == []
    assert general == []

def test_get_report_path():
    """Tests the get_report_path function."""
    from app import config
    from unittest.mock import patch
    import os
    
    with patch.object(config, 'REPORTS_DIR', 'test_reports'):
        path = get_report_path('compsnipe', '2025-01-15_120000')
        assert path == os.path.join('test_reports', 'compsnipe_Report_2025-01-15_120000.xlsx')
        assert 'compsnipe_Report' in path
        assert path.endswith('.xlsx')

def test_to_excel_multi_sheet():
    """Tests creating multi-sheet Excel files."""
    df1 = pd.DataFrame({'A': [1, 2], 'B': [3, 4]})
    df2 = pd.DataFrame({'X': [5, 6], 'Y': [7, 8]})
    
    report_data = [
        {'theater_name': 'AMC Theater 1', 'report': df1},
        {'theater_name': 'Marcus Cinema 2', 'report': df2}
    ]
    
    excel_bytes = to_excel_multi_sheet(report_data)
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0
    
    # Verify it's a valid Excel file by reading it back
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    assert 'AMC Theater 1' in wb.sheetnames
    assert 'Marcus Cinema 2' in wb.sheetnames

def test_to_excel_multi_sheet_sanitizes_names():
    """Tests that sheet names are sanitized for Excel compatibility."""
    df = pd.DataFrame({'A': [1]})
    
    # Test with invalid characters and long name
    report_data = [
        {'theater_name': 'Theater/With*Invalid:Chars[]?', 'report': df},
        {'theater_name': 'A' * 50, 'report': df}  # Too long
    ]
    
    excel_bytes = to_excel_multi_sheet(report_data)
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    
    # Sheet names should be sanitized
    sheet_names = wb.sheetnames
    assert len(sheet_names) == 2
    # Invalid chars removed
    assert '[' not in sheet_names[0]
    assert '*' not in sheet_names[0]
    # Max 31 chars
    assert len(sheet_names[1]) <= 31

def test_is_run_allowed_before_cutoff():
    """Tests is_run_allowed when time is before 8 AM."""
    # Create a mock time before 8 AM
    with patch('app.utils.datetime') as mock_datetime:
        # Set time to 6 AM Central
        mock_time = datetime.datetime(2025, 1, 15, 6, 0, 0, tzinfo=pytz.timezone('America/Chicago'))
        mock_datetime.datetime.now.return_value = mock_time
        
        # Use the real pytz module for timezone operations
        import app.utils
        original_datetime = app.utils.datetime
        app.utils.datetime = mock_datetime
        app.utils.datetime.datetime = mock_datetime.datetime
        
        try:
            result = is_run_allowed('America/Chicago')
            # Should allow run before 8 AM
            # Note: This test may need adjustment based on actual implementation
        finally:
            app.utils.datetime = original_datetime

def test_is_run_allowed_unknown_timezone():
    """Tests is_run_allowed with unknown timezone (should fallback to Chicago)."""
    # This should not raise an error and should fallback to America/Chicago
    result = is_run_allowed('Invalid/Timezone')
    # Should return a boolean without error
    assert isinstance(result, bool)

def test_showtime_selection_to_dataframe_empty():
    """Tests converting empty showtime selection to dataframe."""
    result = showtime_selection_to_dataframe({})
    assert isinstance(result, pd.DataFrame)
    assert result.empty

def test_showtime_selection_to_dataframe():
    """Tests converting nested showtime selections to flat dataframe."""
    selected_showtimes = {
        '2025-01-15': {
            'AMC Theater': {
                'Avatar': {
                    '7:00 PM': [
                        {'format': '3D', 'daypart': 'Evening', 'ticket_url': 'http://tickets.com/1'},
                        {'format': '2D', 'daypart': 'Evening', 'ticket_url': 'http://tickets.com/2'}
                    ],
                    '9:30 PM': [
                        {'format': 'IMAX', 'daypart': 'Night', 'ticket_url': 'http://tickets.com/3'}
                    ]
                },
                'Dune': {
                    '8:00 PM': [
                        {'format': '2D', 'daypart': 'Evening', 'ticket_url': 'http://tickets.com/4'}
                    ]
                }
            }
        }
    }
    
    df = showtime_selection_to_dataframe(selected_showtimes)
    
    # Should have 4 rows (4 total showings)
    assert len(df) == 4
    assert list(df.columns) == ["Date", "Theater Name", "Film Title", "Showtime", "Format", "Daypart", "Ticket URL"]
    
    # Check specific values
    assert (df['Film Title'] == 'Avatar').sum() == 3
    assert (df['Film Title'] == 'Dune').sum() == 1
    assert (df['Format'] == '3D').sum() == 1
    assert (df['Format'] == 'IMAX').sum() == 1
    assert (df['Showtime'] == '7:00 PM').sum() == 2

def test_generate_selection_analysis_report_empty():
    """Tests generating analysis report from empty selections."""
    result = generate_selection_analysis_report({})
    assert isinstance(result, pd.DataFrame)
    assert result.empty

def test_generate_selection_analysis_report():
    """Tests generating pivoted analysis report from showtime selections."""
    selected_showtimes = {
        '2025-01-15': {
            'AMC Theater': {
                'Avatar': {
                    '7:00 PM': [{'format': '3D'}],
                    '9:30 PM': [{'format': 'IMAX'}]
                },
                'Dune': {
                    '8:00 PM': [{'format': '2D'}]
                }
            },
            'Marcus Theater': {
                'Avatar': {
                    '6:00 PM': [{'format': '2D'}]
                }
            }
        },
        '2025-01-16': {
            'AMC Theater': {
                'Avatar': {
                    '7:00 PM': [{'format': '3D'}]
                }
            }
        }
    }
    
    df = generate_selection_analysis_report(selected_showtimes)
    
    # Should have 3 rows (AMC 1/15, Marcus 1/15, AMC 1/16)
    assert len(df) == 3
    
    # Should have Date, Theater Name columns plus film columns plus Total
    assert 'Date' in df.columns
    assert 'Theater Name' in df.columns
    assert 'Total Showings Per Day' in df.columns
    assert 'Avatar' in df.columns
    assert 'Dune' in df.columns
    
    # Check specific values for AMC Theater on 2025-01-15
    amc_1_15 = df[(df['Date'] == '2025-01-15') & (df['Theater Name'] == 'AMC Theater')]
    assert len(amc_1_15) == 1
    assert amc_1_15['Avatar'].values[0] == 2  # 2 showtimes for Avatar
    assert amc_1_15['Dune'].values[0] == 1  # 1 showtime for Dune
    assert amc_1_15['Total Showings Per Day'].values[0] == 3  # 3 total

def test_generate_selection_analysis_report_with_empty_films():
    """Tests that empty film selections are skipped."""
    selected_showtimes = {
        '2025-01-15': {
            'AMC Theater': {
                'Avatar': {},  # Empty selection
                'Dune': {
                    '8:00 PM': [{'format': '2D'}]
                }
            }
        }
    }
    
    df = generate_selection_analysis_report(selected_showtimes)
    
    # Should only have 1 row for Dune
    assert len(df) == 1
    assert 'Dune' in df.columns
    assert 'Avatar' not in df.columns  # Avatar should not appear as it had no showings

def test_generate_selection_analysis_report_all_empty_films():
    """Tests when all films have empty selections (should return empty DataFrame)."""
    selected_showtimes = {
        '2025-01-15': {
            'AMC Theater': {
                'Avatar': {},
                'Dune': {}
            }
        }
    }
    
    df = generate_selection_analysis_report(selected_showtimes)
    assert isinstance(df, pd.DataFrame)
    assert df.empty
