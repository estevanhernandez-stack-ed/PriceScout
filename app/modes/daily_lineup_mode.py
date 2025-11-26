"""
Daily Lineup Mode - Print-ready theater film schedules
Generates printable daily lineups for individual theaters with blank columns for manual theater numbers.
"""

import streamlit as st
import pandas as pd
import re
from datetime import datetime, date, timedelta, time as dt_time
from app import db_adapter as database
from app.utils import run_async_in_thread


def parse_showtime_for_sort(showtime_str):
    """
    Parse showtime string to a time object for proper chronological sorting.

    Handles formats: HH:MM, H:MM, HH:MM:SS, H:MM:SS
    Returns a time object that sorts correctly.
    """
    if not showtime_str:
        return dt_time(23, 59, 59)  # Put empty times at the end

    try:
        showtime_str = str(showtime_str).strip()

        # Try different formats
        for fmt in ['%H:%M:%S', '%H:%M', '%I:%M %p', '%I:%M:%S %p']:
            try:
                return datetime.strptime(showtime_str, fmt).time()
            except ValueError:
                continue

        # Fallback: try to extract hours and minutes manually
        parts = showtime_str.replace(' ', ':').split(':')
        if len(parts) >= 2:
            hour = int(parts[0])
            minute = int(parts[1])
            return dt_time(hour % 24, minute % 60)

        return dt_time(23, 59, 59)  # Default to end of day
    except:
        return dt_time(23, 59, 59)


def parse_runtime_minutes(runtime_str):
    """
    Parse runtime string to minutes.

    Handles formats: "120 min", "2h 30m", "2:30", "120", "2 hr 30 min"
    Returns integer minutes or None if parsing fails.
    """
    if not runtime_str:
        return None

    try:
        runtime_str = str(runtime_str).strip().lower()

        # Try simple number (assume minutes)
        if runtime_str.isdigit():
            return int(runtime_str)

        # Pattern: "120 min" or "120min"
        match = re.match(r'^(\d+)\s*min', runtime_str)
        if match:
            return int(match.group(1))

        # Pattern: "2h 30m" or "2h30m" or "2 h 30 m"
        match = re.match(r'^(\d+)\s*h(?:r|our)?s?\s*(\d+)?\s*m', runtime_str)
        if match:
            hours = int(match.group(1))
            minutes = int(match.group(2)) if match.group(2) else 0
            return hours * 60 + minutes

        # Pattern: "2:30" (hours:minutes)
        match = re.match(r'^(\d+):(\d+)$', runtime_str)
        if match:
            hours = int(match.group(1))
            minutes = int(match.group(2))
            return hours * 60 + minutes

        # Pattern: just hours "2h" or "2 hours"
        match = re.match(r'^(\d+)\s*h(?:r|our)?s?$', runtime_str)
        if match:
            return int(match.group(1)) * 60

        return None
    except:
        return None


def calculate_outtime(showtime_str, runtime_minutes):
    """
    Calculate the end time (outtime) given a showtime and runtime.

    Args:
        showtime_str: Showtime in HH:MM or HH:MM:SS format
        runtime_minutes: Runtime in minutes

    Returns:
        Formatted outtime string or None if calculation fails
    """
    if not showtime_str or not runtime_minutes:
        return None

    try:
        # Parse the showtime
        time_obj = parse_showtime_for_sort(showtime_str)
        if time_obj == dt_time(23, 59, 59):  # Failed to parse
            return None

        # Create a datetime to do the math
        base_date = datetime(2000, 1, 1, time_obj.hour, time_obj.minute, time_obj.second)
        end_datetime = base_date + timedelta(minutes=runtime_minutes)

        # Format the outtime (same format as showtime display)
        return end_datetime.strftime('%I:%M %p').lstrip('0')
    except:
        return None


def compact_film_title(title, remove_year=True, remove_articles=False, max_words=None):
    """
    Make film titles more compact for narrow column display.

    Args:
        title: Original film title
        remove_year: Remove bracketed year like "(2024)" from end
        remove_articles: Remove leading articles (The, A, An)
        max_words: Limit title to first N words (None = no limit)

    Returns:
        Compacted title string
    """
    if not title:
        return title

    result = title.strip()

    # Remove bracketed year at end (e.g., "(2024)", "(2025)")
    if remove_year:
        result = re.sub(r'\s*\(\d{4}\)\s*$', '', result)

    # Optionally remove leading articles
    if remove_articles:
        result = re.sub(r'^(The|A|An)\s+', '', result, flags=re.IGNORECASE)

    # Limit to max words if specified
    if max_words and max_words > 0:
        words = result.split()
        if len(words) > max_words:
            result = ' '.join(words[:max_words])

    return result.strip()


def render_daily_lineup_mode(cache_data, selected_company):
    """
    Main render function for Daily Lineup Mode.
    Allows theater staff to scrape and generate print-ready daily film schedules.
    """
    st.title("📋 Daily Lineup - Theater Schedule")
    st.info(
        "Scrape your theater's showtimes and generate a print-ready daily film lineup. "
        "Perfect for posting schedules or distribution to staff."
    )

    # Get Marcus and Movie Tavern theaters from cache
    all_theaters = []
    if cache_data:
        for market_name, market_data in cache_data.get('markets', {}).items():
            all_theaters.extend(market_data.get('theaters', []))

    # Filter to only Marcus and Movie Tavern theaters
    company_theaters = [t for t in all_theaters if t.get('company') in ['Marcus', 'Movie Tavern']]
    theater_names = sorted([t['name'] for t in company_theaters if 'name' in t])

    if not theater_names:
        st.warning("No Marcus or Movie Tavern theaters found in cache. Please build the theater cache first.")
        return

    # Theater selection
    st.subheader("Select Theater")

    # Check if user has a default theater (future feature)
    default_theater_index = 0
    if 'user_default_theater' in st.session_state:
        try:
            default_theater_index = theater_names.index(st.session_state.user_default_theater)
        except ValueError:
            pass

    selected_theater = st.selectbox(
        "Theater",
        options=theater_names,
        index=default_theater_index,
        help="Select the theater to generate a lineup for"
    )

    # Get theater object
    selected_theater_obj = next((t for t in company_theaters if t['name'] == selected_theater), None)

    if not selected_theater_obj:
        st.error("Theater not found in cache.")
        return

    st.divider()

    # Date picker
    st.subheader("Select Date")

    today = date.today()
    selected_date_obj = st.date_input(
        "Date",
        value=today,
        min_value=today,
        max_value=today + timedelta(days=30),
        help="Select the date to scrape and generate the lineup for"
    )

    selected_date = selected_date_obj.strftime('%Y-%m-%d')

    st.divider()

    # Display options
    st.subheader("Display Options")
    col_opt1, col_opt2, col_opt3, col_opt4 = st.columns(4)

    with col_opt1:
        compact_titles = st.checkbox(
            "Compact Titles",
            value=True,
            help="Remove year from titles (e.g., 'Wicked (2024)' → 'Wicked') for narrower columns"
        )

    with col_opt2:
        remove_articles = st.checkbox(
            "Remove Leading Articles",
            value=False,
            help="Remove 'The', 'A', 'An' from start of titles (e.g., 'The Wild Robot' → 'Wild Robot')"
        )

    with col_opt3:
        max_words = st.selectbox(
            "Max Words",
            options=[0, 2, 3, 4, 5],
            index=2,  # Default to 3 words
            format_func=lambda x: "No limit" if x == 0 else f"{x} words",
            help="Limit titles to first N words (e.g., 'Now You See Me: Now You Don't' → 'Now You See')"
        )

    with col_opt4:
        show_outtime = st.checkbox(
            "Show Out Time",
            value=True,
            help="Calculate and display end time based on film runtime"
        )

    st.divider()

    # Scrape and Generate button
    if st.button("🔄 Get Latest Showtimes & Generate Lineup", type="primary", use_container_width=True):
        scrape_and_generate(selected_theater_obj, selected_theater, selected_date, selected_date_obj,
                          compact_titles=compact_titles, remove_articles=remove_articles,
                          max_words=max_words if max_words > 0 else None, show_outtime=show_outtime)


def scrape_and_generate(theater_obj, theater_name, date_str, date_obj, compact_titles=True, remove_articles=False, max_words=None, show_outtime=True):
    """Scrape showtimes for a single theater for one date and generate lineup"""
    from app.scraper import Scraper

    theater_url = theater_obj.get('url', '')

    if not theater_url:
        st.error(f"No URL found for {theater_name}")
        return

    st.subheader(f"🔄 Scraping {theater_name} for {date_obj.strftime('%A, %B %d, %Y')}")

    with st.spinner(f"Fetching latest showtimes..."):
        # Initialize scraper
        scout = Scraper()

        # Scrape this theater for this date
        thread, result_func = run_async_in_thread(
            scout.get_all_showings_for_theaters,
            [theater_obj],
            date_str
        )
        thread.join()
        status, result, _, _ = result_func()

        if status == 'success' and result:
            # Save to database using upsert_showings
            database.upsert_showings(result, date_str)

            # Count total showings
            total_showings = sum(len(showings) for showings in result.values())
            st.success(f"✅ Successfully scraped {total_showings} showtimes")
        else:
            st.error(f"Failed to scrape showtimes for {date_str}")
            return

    # Generate the lineup
    st.divider()
    generate_daily_lineup(theater_name, date_str, date_obj, compact_titles=compact_titles,
                         remove_articles=remove_articles, max_words=max_words, show_outtime=show_outtime)


def generate_daily_lineup(theater_name, date_str, date_obj, compact_titles=True, remove_articles=False, max_words=None, show_outtime=True):
    """Generate and display the daily lineup"""

    # Query showings for this theater and date using SQLAlchemy
    from app.db_adapter import get_session, Showing, Film, config
    from sqlalchemy.orm import aliased
    from datetime import datetime as dt

    # Convert date_str to date object if needed
    if isinstance(date_str, str):
        play_date = dt.strptime(date_str, '%Y-%m-%d').date()
    else:
        play_date = date_str

    with get_session() as session:
        company_id = getattr(config, 'CURRENT_COMPANY_ID', None)

        # Query showings with optional film runtime for outtime calculation
        query = session.query(
            Showing.film_title,
            Showing.showtime,
            Showing.format,
            Showing.daypart,
            Film.runtime
        ).outerjoin(
            Film,
            (Showing.film_title == Film.film_title) & (Showing.company_id == Film.company_id)
        )

        if company_id:
            query = query.filter(Showing.company_id == company_id)

        query = query.filter(
            Showing.theater_name == theater_name,
            Showing.play_date == play_date
        ).order_by(Showing.showtime, Showing.film_title)

        results = query.all()

        if not results:
            st.warning(f"No showtimes found for {theater_name} on {date_str}")
            return

        # Convert to DataFrame
        df = pd.DataFrame(
            results,
            columns=['film_title', 'showtime', 'format', 'daypart', 'runtime']
        )

    # Sort by showtime properly (not alphabetically)
    # This fixes the issue where "10:00" would appear before "9:00"
    df['_sort_time'] = df['showtime'].apply(parse_showtime_for_sort)
    df = df.sort_values(by=['_sort_time', 'film_title']).drop(columns=['_sort_time'])

    # Process the data - create one row per showtime (chronological)
    lineup_data = []

    for _, row in df.iterrows():
        # Format showtime (remove seconds if present)
        formatted_time = format_showtime(row['showtime'])

        # Calculate outtime if runtime is available and option is enabled
        outtime = None
        if show_outtime and row.get('runtime'):
            runtime_mins = parse_runtime_minutes(row['runtime'])
            if runtime_mins:
                outtime = calculate_outtime(row['showtime'], runtime_mins)

        # Get format indicator for this specific showing
        format_indicator = get_format_indicators([row['format']])

        # Apply title compacting if enabled
        film_title = row['film_title']
        if compact_titles or remove_articles or max_words:
            film_title = compact_film_title(film_title, remove_year=compact_titles, remove_articles=remove_articles, max_words=max_words)

        row_data = {
            'Theater #': '',  # Blank column for manual entry
            'Showtime': formatted_time,
            'Film Title': film_title,
            'Format': format_indicator
        }

        # Add outtime column if enabled
        if show_outtime:
            row_data['Out Time'] = outtime if outtime else ''

        lineup_data.append(row_data)

    # Create DataFrame for display
    lineup_df = pd.DataFrame(lineup_data)

    # Display header
    st.success(f"✅ Daily Lineup Generated for {theater_name}")
    st.subheader(f"{date_obj.strftime('%A, %B %d, %Y')}")

    # Build column config
    column_config = {
        'Theater #': st.column_config.TextColumn(
            'Theater #',
            width='small',
            help='Leave blank for manual entry'
        ),
        'Showtime': st.column_config.TextColumn(
            'Showtime',
            width='small'
        ),
        'Film Title': st.column_config.TextColumn(
            'Film Title',
            width='large'
        ),
        'Format': st.column_config.TextColumn(
            'Format',
            width='medium',
            help='3D, IMAX, PLF indicators'
        )
    }

    # Add Out Time column config if showing outtime
    if show_outtime:
        column_config['Out Time'] = st.column_config.TextColumn(
            'Out Time',
            width='small',
            help='Calculated end time based on film runtime'
        )

    # Display the lineup table
    st.dataframe(
        lineup_df,
        use_container_width=True,
        hide_index=True,
        column_config=column_config
    )

    # Print instructions
    st.divider()
    st.info(
        "📌 **Printing Instructions:**\n"
        "1. Use your browser's print function (Ctrl+P or Cmd+P)\n"
        "2. Set orientation to 'Landscape' for best results\n"
        "3. Adjust scale if needed to fit all content on one page\n"
        "4. The 'Theater #' column can be filled in by hand after printing"
    )

    # Download options
    st.subheader("Download Options")
    col1, col2 = st.columns(2)

    with col1:
        # CSV download
        csv_data = lineup_df.to_csv(index=False)
        st.download_button(
            label="📄 Download as CSV",
            data=csv_data,
            file_name=f"daily_lineup_{theater_name}_{date_str}.csv",
            mime="text/csv",
            use_container_width=True
        )

    with col2:
        # Excel download with enhanced formatting for easy editing
        try:
            from io import BytesIO
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                lineup_df.to_excel(writer, index=False, sheet_name='Daily Lineup', startrow=2)

                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Daily Lineup']

                # Add theater name and date header
                worksheet['A1'] = theater_name
                worksheet['A1'].font = Font(size=14, bold=True)
                worksheet['A2'] = date_obj.strftime('%A, %B %d, %Y')
                worksheet['A2'].font = Font(size=12, bold=True)

                # Format header row (row 3)
                header_fill = PatternFill(start_color='8b0e04', end_color='8b0e04', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True, size=11)

                for col_num, column in enumerate(lineup_df.columns, 1):
                    cell = worksheet.cell(row=3, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Set column widths for easy editing
                worksheet.column_dimensions['A'].width = 12  # Theater #
                worksheet.column_dimensions['B'].width = 45  # Film Title
                worksheet.column_dimensions['C'].width = 50  # Showtimes
                worksheet.column_dimensions['D'].width = 20  # Format

                # Add borders to all cells with data
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for row in worksheet.iter_rows(min_row=3, max_row=len(lineup_df) + 3, min_col=1, max_col=4):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical='top', wrap_text=True)

                # Alternate row colors for easier reading
                light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                for row_num in range(4, len(lineup_df) + 4, 2):  # Every other row
                    for col_num in range(1, 5):
                        worksheet.cell(row=row_num, column=col_num).fill = light_fill

            excel_data = output.getvalue()

            st.download_button(
                label="📊 Download as Excel (Editable)",
                data=excel_data,
                file_name=f"daily_lineup_{theater_name}_{date_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        except ImportError:
            st.caption("Excel export requires openpyxl package")

    # Summary statistics
    st.divider()
    col1, col2, col3 = st.columns(3)
    with col1:
        unique_films = lineup_df['Film Title'].nunique()
        st.metric("Total Films", unique_films)
    with col2:
        total_showtimes = len(lineup_df)
        st.metric("Total Showtimes", total_showtimes)
    with col3:
        # Count premium formats
        premium_count = sum(1 for _, row in lineup_df.iterrows() if row['Format'] and row['Format'] != 'Standard')
        st.metric("Premium Format Shows", premium_count)


def format_showtime(showtime_str):
    """Format showtime string to be more readable"""
    try:
        # Handle different time formats
        if len(showtime_str) == 5:  # HH:MM format
            time_obj = datetime.strptime(showtime_str, '%H:%M')
        elif len(showtime_str) == 8:  # HH:MM:SS format
            time_obj = datetime.strptime(showtime_str, '%H:%M:%S')
        else:
            return showtime_str

        # Convert to 12-hour format with AM/PM
        return time_obj.strftime('%I:%M %p').lstrip('0')
    except:
        return showtime_str


def get_format_indicators(formats):
    """Convert format codes to readable indicators"""
    indicators = []

    for fmt in formats:
        # Handle None, empty, or whitespace-only values
        if not fmt or (isinstance(fmt, str) and not fmt.strip()):
            continue

        fmt_str = str(fmt).strip()
        fmt_upper = fmt_str.upper()

        # Skip standard 2D formats
        if fmt_upper in ['STANDARD', '2D', 'STANDARD 2D']:
            continue

        # Common format mappings
        if '3D' in fmt_upper:
            indicators.append('3D')
        if 'IMAX' in fmt_upper:
            indicators.append('IMAX')
        if 'ULTRASCREEN' in fmt_upper:
            indicators.append('UltraScreen')
        if 'PLF' in fmt_upper or 'SUPERSCREEN' in fmt_upper or 'PREMIUM' in fmt_upper:
            indicators.append('PLF')
        if 'DFX' in fmt_upper:
            indicators.append('DFX')
        if 'DOLBY' in fmt_upper:
            indicators.append('Dolby')
        if 'XD' in fmt_upper:
            indicators.append('XD')
        if 'RPX' in fmt_upper:
            indicators.append('RPX')
        if 'DBOX' in fmt_upper or 'D-BOX' in fmt_upper:
            indicators.append('D-BOX')

        # If no specific format matched but it's not standard/2D, show the original
        if not indicators and fmt_upper not in ['STANDARD', '2D', 'STANDARD 2D']:
            indicators.append(fmt_str)

    if not indicators:
        return 'Standard'

    # Remove duplicates and join
    return ', '.join(sorted(set(indicators)))
